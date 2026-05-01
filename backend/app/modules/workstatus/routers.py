"""
근태 라우터
─────────────────────────────────────────
1. 레거시 (username + password)   : POST /check-in, /break-start, /break-end, /check-out
2. 키오스크 (system JWT + user_id): GET /employees, GET /today/{user_id},
                                    POST /kiosk/check-in, /kiosk/break-start,
                                         /kiosk/break-end, /kiosk/check-out
3. 관리자                         : GET /admin/monthly, GET /admin/template,
                                    POST /admin/bulk-import
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime, time, timedelta
from typing import List, Optional

import openpyxl
from openpyxl.styles import numbers
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_admin, get_current_user
from app.modules.auth.models import PositionEnum, StatusEnum, User
from app.modules.auth.services import verify_password
from app.modules.workstatus import models, schemas
from app.modules.workstatus.models import AttendanceEvent, EventType
from app.modules.workstatus.services import AttendanceService
from app.utils.permission_utils import is_system

router = APIRouter()

_KIOSK_POSITIONS = {PositionEnum.crew, PositionEnum.leader}


# ════════════════════════════════════════════════════════
# 공통 유틸
# ════════════════════════════════════════════════════════

def _get_active_work_date(db: Session, user_id: int) -> date:
    """
    야간 근무 지원: 오늘 CLOCK_IN이 없으면 전날(어제) 확인.
    어제 CLOCK_IN은 있고 CLOCK_OUT이 없는 경우 → 야간 근무 중으로 판단 → 어제 날짜 반환.
    그 외에는 오늘 날짜 반환.
    """
    today = datetime.now().date()
    events_today = AttendanceService.get_events_for_date(db, user_id, today)
    if EventType.CLOCK_IN in events_today:
        return today

    yesterday = today - timedelta(days=1)
    events_yesterday = AttendanceService.get_events_for_date(db, user_id, yesterday)
    if (
        EventType.CLOCK_IN in events_yesterday
        and EventType.CLOCK_OUT not in events_yesterday
    ):
        return yesterday

    return today


# ════════════════════════════════════════════════════════
# Excel 파싱 유틸리티 — serial number / 다양한 포맷 대응
# ════════════════════════════════════════════════════════

# Excel serial number 기준일 (1900-01-01 = 1, 단 1900-02-29 버그로 +2)
_EXCEL_EPOCH = datetime(1899, 12, 30)


def _parse_excel_date(val) -> date:
    """
    Excel 셀 값 → date 변환.
    지원 형식:
      - datetime / date 객체 (openpyxl이 Date 서식 셀에서 반환)
      - int / float (Excel serial number, 예: 46082 → 2026-03-01)
      - str: YYYY-MM-DD, YYYY/MM/DD, YYYY.MM.DD
    """
    if val is None:
        raise ValueError("날짜 값이 없습니다")

    # datetime / date 객체
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val

    # 숫자형 → Excel serial number
    if isinstance(val, (int, float)):
        serial = int(val)
        if serial < 1:
            raise ValueError(f"유효하지 않은 날짜 serial number: {val}")
        return (_EXCEL_EPOCH + timedelta(days=serial)).date()

    # 문자열 파싱
    s = str(val).strip()
    if not s:
        raise ValueError("날짜 값이 비어 있습니다")

    # 수식 결과 감지 (=로 시작하는 문자열)
    if s.startswith("="):
        raise ValueError(f"수식은 지원되지 않습니다. 값으로 변환 후 업로드하세요: {s}")

    # 구분자 통일: / . → -
    normalized = re.sub(r"[/.]", "-", s)

    # ISO 형식 시도 (YYYY-MM-DD)
    try:
        return date.fromisoformat(normalized)
    except ValueError:
        pass

    # 추가 포맷 시도
    for fmt in ("%Y-%m-%d", "%m-%d-%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(normalized, fmt).date()
        except ValueError:
            pass

    raise ValueError(f"인식할 수 없는 날짜 형식: {val}")


def _parse_excel_time(val) -> Optional[time]:
    """
    Excel 셀 값 → time 변환.
    지원 형식:
      - None → None 반환
      - time 객체 (openpyxl이 Time 서식 셀에서 반환)
      - datetime 객체 → .time() 추출
      - float (Excel serial number, 예: 0.375 → 09:00, 0.75 → 18:00)
      - str: HH:MM, HH:MM:SS, H:MM AM/PM
    """
    if val is None:
        return None

    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()

    # 숫자형 → Excel time serial (하루 = 1.0)
    if isinstance(val, (int, float)):
        fval = float(val)
        # 0.0 ~ 1.0 범위: 시간 serial
        if 0.0 <= fval < 1.0:
            total_seconds = int(round(fval * 86400))
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            return time(hours, minutes, seconds)
        # 1 이상이면 datetime serial (날짜+시간)
        if fval >= 1.0:
            dt = _EXCEL_EPOCH + timedelta(days=fval)
            return dt.time()
        raise ValueError(f"유효하지 않은 시간 serial number: {val}")

    # 문자열 파싱
    s = str(val).strip()
    if not s or s == "-":
        return None

    # 수식 감지
    if s.startswith("="):
        raise ValueError(f"수식은 지원되지 않습니다. 값으로 변환 후 업로드하세요: {s}")

    # 다양한 포맷 시도
    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            pass

    # 마지막 시도: 숫자 문자열 (소수점 포함)
    try:
        fval = float(s)
        if 0.0 <= fval < 1.0:
            total_seconds = int(round(fval * 86400))
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            return time(hours, minutes, seconds)
    except ValueError:
        pass

    raise ValueError(f"인식할 수 없는 시간 형식: {val}")


def require_system_user(user: User = Depends(get_current_user)) -> User:
    if not is_system(user):
        raise HTTPException(status_code=403, detail="시스템 계정만 접근 가능합니다.")
    return user


def authenticate_attendance_user(db: Session, username: str, password: str) -> User:
    user = db.execute(
        select(User).where(User.username == username)
    ).scalar_one_or_none()

    if not user or not verify_password(password, user.password):
        raise HTTPException(status_code=401, detail="아이디 또는 비밀번호가 올바르지 않습니다.")
    if not user.is_active:
        raise HTTPException(status_code=400, detail="비활성화된 계정입니다.")
    return user


def _get_kiosk_user(db: Session, user_id: int) -> User:
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="직원을 찾을 수 없습니다.")
    if user.status != StatusEnum.approved:
        raise HTTPException(status_code=400, detail="승인된 계정이 아닙니다.")
    if user.position not in _KIOSK_POSITIONS:
        raise HTTPException(status_code=400, detail="근태 대상 직급이 아닙니다.")
    return user


def _add_event(
    db: Session,
    user_id: int,
    work_date: date,
    event_type: EventType,
    event_time: Optional[time] = None,
    note: Optional[str] = None,
) -> AttendanceEvent:
    """이벤트 추가 (중복 시 409)"""
    if event_time is None:
        event_time = datetime.now().time()

    ev = AttendanceEvent(
        user_id=user_id,
        work_date=work_date,
        event_type=event_type,
        event_time=event_time,
        note=note,
    )
    db.add(ev)
    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail=f"이미 {event_type.value} 기록이 있습니다.",
        )
    return ev


def _build_summary(db: Session, user_id: int, work_date: date) -> schemas.DailySummary:
    summary = AttendanceService.get_today_summary(db, user_id, work_date)
    user = db.get(User, user_id)
    if summary is None:
        return schemas.DailySummary(
            user_id=user_id,
            user_name=user.name if user else None,
            work_date=work_date,
        )
    from app.modules.workstatus.services import _check_break_warning
    events = AttendanceService.get_events_for_date(db, user_id, work_date)
    dm, nm, bm = AttendanceService.calc_work_minutes(events, work_date)
    clock_in_ev = events.get(EventType.CLOCK_IN)
    return schemas.DailySummary(
        user_id=user_id,
        user_name=user.name if user else None,
        position=user.position.value if user else None,
        work_date=work_date,
        check_in=summary["check_in"],
        break_start=summary["break_start"],
        break_end=summary["break_end"],
        check_out=summary["check_out"],
        total_work_hours=float(
            AttendanceService.minutes_to_hours(dm + nm)
        ),
        day_hours=float(AttendanceService.minutes_to_hours(dm)),
        night_hours=float(AttendanceService.minutes_to_hours(nm)),
        break_warning=_check_break_warning(dm + nm, bm),
        note=clock_in_ev.note if clock_in_ev else None,
    )


# ════════════════════════════════════════════════════════
# 레거시 API (username + password)
# ════════════════════════════════════════════════════════

class AttendanceAuthInput(BaseModel):
    username: str
    password: str


@router.post("/check-in", response_model=schemas.DailySummary, summary="[레거시] 출근")
def check_in(
    payload: AttendanceAuthInput,
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    user = authenticate_attendance_user(db, payload.username, payload.password)
    today = datetime.now().date()

    _add_event(db, user.id, today, EventType.CLOCK_IN)
    db.commit()
    return _build_summary(db, user.id, today)


@router.post("/break-start", response_model=schemas.DailySummary, summary="[레거시] 휴식 시작")
def break_start(
    payload: AttendanceAuthInput,
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    user = authenticate_attendance_user(db, payload.username, payload.password)
    work_date = _get_active_work_date(db, user.id)

    events = AttendanceService.get_events_for_date(db, user.id, work_date)
    if EventType.CLOCK_IN not in events:
        raise HTTPException(status_code=400, detail="출근 먼저 해주세요.")
    if EventType.CLOCK_OUT in events:
        raise HTTPException(status_code=400, detail="이미 퇴근한 기록이 있습니다.")

    _add_event(db, user.id, work_date, EventType.BREAK_START)
    db.commit()
    return _build_summary(db, user.id, work_date)


@router.post("/break-end", response_model=schemas.DailySummary, summary="[레거시] 복귀")
def break_end(
    payload: AttendanceAuthInput,
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    user = authenticate_attendance_user(db, payload.username, payload.password)
    work_date = _get_active_work_date(db, user.id)

    events = AttendanceService.get_events_for_date(db, user.id, work_date)
    if EventType.BREAK_START not in events:
        raise HTTPException(status_code=400, detail="휴식 시작 기록이 없습니다.")

    _add_event(db, user.id, work_date, EventType.BREAK_END)
    db.commit()
    return _build_summary(db, user.id, work_date)


@router.post("/check-out", response_model=schemas.DailySummary, summary="[레거시] 퇴근")
def check_out(
    payload: AttendanceAuthInput,
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    user = authenticate_attendance_user(db, payload.username, payload.password)
    work_date = _get_active_work_date(db, user.id)

    events = AttendanceService.get_events_for_date(db, user.id, work_date)
    if EventType.CLOCK_IN not in events:
        raise HTTPException(status_code=400, detail="출근 기록이 없습니다.")
    if EventType.BREAK_START in events and EventType.BREAK_END not in events:
        raise HTTPException(status_code=400, detail="휴식 중에는 퇴근할 수 없습니다.")

    _add_event(db, user.id, work_date, EventType.CLOCK_OUT)
    AttendanceService.handle_check_out(db, user.id, work_date)
    db.commit()
    return _build_summary(db, user.id, work_date)


# ════════════════════════════════════════════════════════
# 키오스크 API (system JWT + user_id)
# ════════════════════════════════════════════════════════

@router.get(
    "/employees",
    response_model=schemas.KioskEmployeesResponse,
    summary="[키오스크] 직원 목록",
)
def get_kiosk_employees(
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    users = (
        db.query(User)
        .filter(
            User.status == StatusEnum.approved,
            User.is_active == 1,
            User.position.in_(_KIOSK_POSITIONS),
        )
        .order_by(User.name)
        .all()
    )
    return schemas.KioskEmployeesResponse(
        items=[
            schemas.KioskEmployeeDTO(
                id=u.id,
                name=u.name,
                position=u.position.value,
                username=u.username,
                profile_image=u.profile_image,
            )
            for u in users
        ]
    )


@router.get(
    "/today/{user_id}",
    response_model=Optional[schemas.DailySummary],
    summary="[키오스크] 오늘 근태 조회",
)
def get_today_kiosk(
    user_id: int,
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    # 야간 근무 지원: 현재 열려있는 근무 날짜 반환
    work_date = _get_active_work_date(db, user_id)
    summary = AttendanceService.get_today_summary(db, user_id, work_date)
    if not summary:
        return None
    return _build_summary(db, user_id, work_date)


@router.post(
    "/kiosk/check-in",
    response_model=schemas.DailySummary,
    summary="[키오스크] 출근",
)
def kiosk_check_in(
    payload: schemas.KioskActionInput,
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    user = _get_kiosk_user(db, payload.user_id)
    today = datetime.now().date()

    _add_event(db, user.id, today, EventType.CLOCK_IN)
    db.commit()
    return _build_summary(db, user.id, today)


@router.post(
    "/kiosk/break-start",
    response_model=schemas.DailySummary,
    summary="[키오스크] 휴식 시작",
)
def kiosk_break_start(
    payload: schemas.KioskActionInput,
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    user = _get_kiosk_user(db, payload.user_id)
    work_date = _get_active_work_date(db, user.id)

    events = AttendanceService.get_events_for_date(db, user.id, work_date)
    if EventType.CLOCK_IN not in events:
        raise HTTPException(status_code=400, detail="출근 먼저 해주세요.")
    if EventType.CLOCK_OUT in events:
        raise HTTPException(status_code=400, detail="이미 퇴근한 기록이 있습니다.")
    if EventType.BREAK_START in events and EventType.BREAK_END not in events:
        raise HTTPException(status_code=400, detail="이미 휴식 중입니다.")

    _add_event(db, user.id, work_date, EventType.BREAK_START)
    db.commit()
    return _build_summary(db, user.id, work_date)


@router.post(
    "/kiosk/break-end",
    response_model=schemas.DailySummary,
    summary="[키오스크] 복귀",
)
def kiosk_break_end(
    payload: schemas.KioskActionInput,
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    user = _get_kiosk_user(db, payload.user_id)
    work_date = _get_active_work_date(db, user.id)

    events = AttendanceService.get_events_for_date(db, user.id, work_date)
    if EventType.BREAK_START not in events:
        raise HTTPException(status_code=400, detail="휴식 시작 기록이 없습니다.")
    if EventType.BREAK_END in events:
        raise HTTPException(status_code=400, detail="이미 복귀한 기록이 있습니다.")

    _add_event(db, user.id, work_date, EventType.BREAK_END)
    db.commit()
    return _build_summary(db, user.id, work_date)


@router.post(
    "/kiosk/check-out",
    response_model=schemas.DailySummary,
    summary="[키오스크] 퇴근",
)
def kiosk_check_out(
    payload: schemas.KioskActionInput,
    db: Session = Depends(get_db),
    _sys: User = Depends(require_system_user),
):
    user = _get_kiosk_user(db, payload.user_id)
    work_date = _get_active_work_date(db, user.id)

    events = AttendanceService.get_events_for_date(db, user.id, work_date)
    if EventType.CLOCK_IN not in events:
        raise HTTPException(status_code=400, detail="출근 기록이 없습니다.")
    if EventType.BREAK_START in events and EventType.BREAK_END not in events:
        raise HTTPException(status_code=400, detail="휴식 중에는 퇴근할 수 없습니다.")
    if EventType.CLOCK_OUT in events:
        raise HTTPException(status_code=400, detail="이미 퇴근 기록이 있습니다.")

    _add_event(db, user.id, work_date, EventType.CLOCK_OUT)
    AttendanceService.handle_check_out(db, user.id, work_date)
    db.commit()
    return _build_summary(db, user.id, work_date)


# ════════════════════════════════════════════════════════
# 관리자 근태 관리 API
# ════════════════════════════════════════════════════════

@router.get(
    "/admin/monthly",
    response_model=schemas.MonthlyAttendanceResponse,
    summary="[관리자] 월별 근태 조회",
)
def admin_monthly_attendance(
    year: int = Query(...),
    month: int = Query(...),
    user_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    records = AttendanceService.get_monthly_attendance(db, year, month, user_id)
    return schemas.MonthlyAttendanceResponse(
        records=[schemas.DailySummary(**r) for r in records],
        total=len(records),
    )


@router.get(
    "/admin/template",
    summary="[관리자] 근태 엑셀 양식 다운로드",
)
def download_attendance_template(
    _admin=Depends(get_current_admin),
):
    """근태 대량 업로드용 엑셀 양식 다운로드"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "근태등록"

    headers = [
        "user_id",
        "name",
        "date",
        "clock_in",
        "break_start",
        "break_end",
        "clock_out",
    ]
    ws.append(headers)

    # 예시 데이터
    ws.append([1, "홍길동", "2026-03-01", "09:00", "12:00", "13:00", "18:00"])

    # ── 셀 서식을 텍스트(@)로 고정 → Excel 자동 날짜/시간 변환 방지 ──
    text_format = numbers.FORMAT_TEXT  # '@'
    for col_idx in range(1, len(headers) + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        # 전체 열에 텍스트 서식 적용 (최대 100행까지 선제 적용)
        for row_idx in range(1, 101):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = text_format
        ws.column_dimensions[col_letter].width = 15

    # 헤더 스타일
    header_fill = PatternFill(start_color="1A0F3C", end_color="1A0F3C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # 안내 사항 시트 추가
    ws_guide = wb.create_sheet("입력안내")
    ws_guide.append(["컬럼", "필수여부", "입력 형식", "예시"])
    ws_guide.append(["user_id", "필수", "직원 ID (숫자)", "1"])
    ws_guide.append(["name", "선택", "직원 이름 (확인용)", "홍길동"])
    ws_guide.append(["date", "필수", "YYYY-MM-DD (텍스트)", "2026-03-01"])
    ws_guide.append(["clock_in", "필수", "HH:MM (텍스트)", "09:00"])
    ws_guide.append(["break_start", "선택", "HH:MM (텍스트)", "12:00"])
    ws_guide.append(["break_end", "선택", "HH:MM (텍스트)", "13:00"])
    ws_guide.append(["clock_out", "필수", "HH:MM (텍스트)", "18:00"])
    ws_guide.append([])
    ws_guide.append(["주의사항:"])
    ws_guide.append(["- 셀 서식을 날짜/시간으로 변경하지 마세요. 텍스트(@) 서식을 유지해주세요."])
    ws_guide.append(["- 같은 날짜의 기존 기록은 덮어씌워집니다."])
    ws_guide.append(["- 수식(=TODAY() 등)은 사용하지 마세요."])

    for col in ws_guide.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_guide.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=attendance_template.xlsx"},
    )


@router.get(
    "/admin/export",
    summary="[관리자] 월별 근태 엑셀 다운로드",
)
def export_attendance_excel(
    year: int = Query(...),
    month: int = Query(...),
    db: Session = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    """월별 전직원 근태 데이터를 엑셀로 다운로드"""
    from collections import defaultdict

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    records = AttendanceService.get_monthly_attendance(db, year, month, None)

    # 직원별로 그룹화
    grouped: dict = defaultdict(list)
    for r in records:
        grouped[r["user_id"]].append(r)

    # 전월 말일이 이번 달 1주차에 포함되는 경우 해당 날짜 레코드 미리 조회
    month_start_date = date(year, month, 1)
    iso_week_start = month_start_date - timedelta(days=month_start_date.weekday())
    prev_overlap_by_user: dict = defaultdict(list)
    if iso_week_start < month_start_date:
        prev_year = iso_week_start.year
        prev_month = iso_week_start.month
        prev_records = AttendanceService.get_monthly_attendance(
            db, prev_year, prev_month, None
        )
        for r in prev_records:
            if r["work_date"] >= iso_week_start and r.get("check_out"):
                prev_overlap_by_user[r["user_id"]].append(r)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}년{month}월 근태"

    headers = ["직원ID", "이름", "날짜", "출근", "휴식시작", "휴식종료", "퇴근", "총근무(h)", "주간(h)", "야간(h)", "주휴(h)", "휴게경고", "비고"]
    ws.append(headers)

    header_fill   = PatternFill(start_color="1A0F3C", end_color="1A0F3C", fill_type="solid")
    header_font   = Font(color="FFFFFF", bold=True, size=11)
    subtotal_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    subtotal_font = Font(bold=True, size=10)
    week_fill     = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    week_font     = Font(color="4338CA", size=9)
    warning_font  = Font(color="DC2626", bold=True, size=10)
    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for user_id, user_records in grouped.items():
        for r in user_records:
            warning = r.get("break_warning") or ""
            ws.append([
                r["user_id"],
                r["user_name"] or "",
                str(r["work_date"]),
                str(r["check_in"])[:5] if r["check_in"] else "",
                str(r["break_start"])[:5] if r["break_start"] else "",
                str(r["break_end"])[:5] if r["break_end"] else "",
                str(r["check_out"])[:5] if r["check_out"] else "",
                round(r["total_work_hours"] or 0, 2),
                round(r["day_hours"] or 0, 2),
                round(r["night_hours"] or 0, 2),
                "",  # 주휴(h) — 데이터 행은 비움
                warning,
                r.get("note") or "",
            ])
            if warning:
                ws.cell(row=ws.max_row, column=12).font = warning_font

        # ── 직원별 소계 행 ────────────────────────────────
        days = len(user_records)
        total_h = round(sum(r["total_work_hours"] or 0 for r in user_records), 2)
        day_h = round(sum(r["day_hours"] or 0 for r in user_records), 2)
        night_h = round(sum(r["night_hours"] or 0 for r in user_records), 2)
        name = user_records[0]["user_name"] or ""

        # 주휴수당 주별 계산 (서비스 로직과 동일한 기준)
        # 전월 겹치는 날짜 포함
        from decimal import Decimal, ROUND_HALF_UP as _HALF_UP
        all_week_records = user_records + prev_overlap_by_user.get(user_id, [])
        work_dates = [r["work_date"] for r in all_week_records if r.get("check_out")]
        weeks_map: dict = {}
        for d in work_dates:
            iso_y, iso_w, _ = d.isocalendar()
            weeks_map.setdefault((iso_y, iso_w), []).append(d)

        week_details = []
        total_weekly_h = Decimal("0.00")
        for (iso_y, iso_w), dates in sorted(weeks_map.items()):
            week_start = min(dates)
            week_end = week_start + timedelta(days=6)
            # 주 종료일이 해당 월 안에 있어야 주휴 발생
            if week_end.month != month:
                continue
            week_total_h = sum(
                (r["total_work_hours"] or 0)
                for r in all_week_records
                if r["work_date"] in dates
            )
            week_total_dec = Decimal(str(round(week_total_h, 2)))
            if week_total_dec >= Decimal("15.00"):
                allowance = min(
                    (week_total_dec / Decimal(40) * Decimal(8)).quantize(
                        Decimal("0.01"), rounding=_HALF_UP
                    ),
                    Decimal("8.00"),
                )
                total_weekly_h += allowance
                note = f"주휴 {float(allowance):.2f}h 발생"
            else:
                allowance = Decimal("0.00")
                note = f"미발생 (주 근무 {float(week_total_dec):.1f}h < 15h)"
            week_details.append((week_start, week_end, week_total_dec, allowance, note))

        subtotal_label = f"[소계] {name} — 근무 {days}일"
        subtotal_row = [subtotal_label, "", "", "", "", "", "", total_h, day_h, night_h, float(total_weekly_h), "", ""]
        ws.append(subtotal_row)

        subtotal_row_idx = ws.max_row
        for col_idx, cell in enumerate(ws[subtotal_row_idx], 1):
            cell.fill = subtotal_fill
            cell.font = subtotal_font
            cell.border = border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left")
            elif col_idx >= 8:
                cell.alignment = Alignment(horizontal="right")

        # ── 주별 명세 행 ──────────────────────────────────
        for week_start, week_end, week_total_dec, allowance, note in week_details:
            label = (
                f"  └ {week_start.strftime('%m/%d')}~{week_end.strftime('%m/%d')}"
                f"  근무 {float(week_total_dec):.1f}h → {note}"
            )
            ws.append([label, "", "", "", "", "", "", "", "", "", "", "", ""])
            week_row_idx = ws.max_row
            for col_idx, cell in enumerate(ws[week_row_idx], 1):
                cell.fill = week_fill
                cell.font = week_font
                cell.border = border
                cell.alignment = Alignment(horizontal="left")

    col_widths = [10, 12, 14, 10, 10, 10, 10, 12, 10, 10, 10, 18, 24]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # A열(소계·주별 명세 레이블)은 넓게
    ws.column_dimensions["A"].width = 48

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"attendance_{year}_{str(month).zfill(2)}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post(
    "/admin/bulk-import",
    response_model=schemas.BulkImportResult,
    summary="[관리자] 근태 엑셀 대량 등록",
)
def bulk_import_attendance(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    """
    엑셀 파일 업로드 → 근태 이벤트 자동 등록
    컬럼: user_id, name, date, clock_in, break_start, break_end, clock_out

    지원하는 데이터 형식:
      - 날짜: YYYY-MM-DD, YYYY/MM/DD, YYYY.MM.DD, datetime 객체, Excel serial number
      - 시간: HH:MM, HH:MM:SS, H:MM AM/PM, time 객체, Excel serial number (0.0~1.0)
      - null/공백: 선택 컬럼(break_start, break_end)은 건너뜀
    """
    try:
        content = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"엑셀 파일 파싱 오류: {e}")

    success_count = 0
    error_count = 0
    errors: List[str] = []
    # 성공적으로 처리된 (user_id, year, month) 조합 수집 — Payroll 재계산에 사용
    affected_payrolls: set[tuple[int, int, int]] = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        # 행 단위 savepoint로 개별 행 에러 시 해당 행만 롤백
        savepoint = db.begin_nested()
        try:
            if len(row) < 7:
                raise ValueError(f"컬럼 수 부족: {len(row)}개 (최소 7개 필요)")

            user_id_raw, name_raw, date_raw, clock_in_raw, break_start_raw, break_end_raw, clock_out_raw = row[:7]

            # 필수값 검증 — 공백 문자열도 누락으로 처리
            def _is_empty(v) -> bool:
                return v is None or (isinstance(v, str) and not v.strip())

            if _is_empty(user_id_raw):
                raise ValueError("필수 컬럼 'user_id' 누락")
            if _is_empty(date_raw):
                raise ValueError("필수 컬럼 'date' 누락")
            if _is_empty(clock_in_raw):
                raise ValueError("필수 컬럼 'clock_in' 누락")
            if _is_empty(clock_out_raw):
                raise ValueError("필수 컬럼 'clock_out' 누락")

            user_id = int(user_id_raw)

            # 날짜 파싱 (serial number, 다양한 포맷 지원)
            work_date = _parse_excel_date(date_raw)

            # 시간 파싱 (serial number, AM/PM 등 지원)
            clock_in_time = _parse_excel_time(clock_in_raw)
            break_start_time = _parse_excel_time(break_start_raw)
            break_end_time = _parse_excel_time(break_end_raw)
            clock_out_time = _parse_excel_time(clock_out_raw)

            # 비즈니스 validation
            if clock_in_time is None:
                raise ValueError("출근 시간이 유효하지 않습니다")
            if clock_out_time is None:
                raise ValueError("퇴근 시간이 유효하지 않습니다")
            if break_start_time and not break_end_time:
                raise ValueError("휴식 시작이 있으면 휴식 종료도 필요합니다")
            if break_end_time and not break_start_time:
                raise ValueError("휴식 종료가 있으면 휴식 시작도 필요합니다")

            # 직원 존재 확인
            user = db.get(User, user_id)
            if not user:
                raise ValueError(f"user_id={user_id} 직원을 찾을 수 없습니다")

            # 기존 이벤트 삭제 (덮어쓰기)
            db.query(AttendanceEvent).filter_by(
                user_id=user_id, work_date=work_date
            ).delete()

            # 이벤트 등록
            if clock_in_time:
                db.add(AttendanceEvent(
                    user_id=user_id, work_date=work_date,
                    event_type=EventType.CLOCK_IN, event_time=clock_in_time,
                ))
            if break_start_time:
                db.add(AttendanceEvent(
                    user_id=user_id, work_date=work_date,
                    event_type=EventType.BREAK_START, event_time=break_start_time,
                ))
            if break_end_time:
                db.add(AttendanceEvent(
                    user_id=user_id, work_date=work_date,
                    event_type=EventType.BREAK_END, event_time=break_end_time,
                ))
            if clock_out_time:
                db.add(AttendanceEvent(
                    user_id=user_id, work_date=work_date,
                    event_type=EventType.CLOCK_OUT, event_time=clock_out_time,
                ))

            db.flush()

            # 이 행에 출근+퇴근이 모두 있으면 Payroll 재계산 대상으로 등록
            if clock_in_time and clock_out_time:
                affected_payrolls.add((user_id, work_date.year, work_date.month))

            success_count += 1

        except Exception as e:
            error_count += 1
            errors.append(f"Row {row_idx}: {e}")
            savepoint.rollback()
            # 오류 행만 롤백, 다른 행은 유지

    # 모든 행 처리 후 영향받은 (user_id, year, month) 별로 Payroll 전체 재계산
    for payroll_user_id, year, month in affected_payrolls:
        AttendanceService.recalculate_payroll_for_month(
            db, payroll_user_id, year, month
        )

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"DB 저장 오류: {e}")

    return schemas.BulkImportResult(
        success_count=success_count,
        error_count=error_count,
        errors=errors,
    )


@router.post(
    "/admin/record",
    response_model=schemas.DailySummary,
    summary="[관리자] 근태 생성/덮어쓰기",
)
def admin_create_record(
    payload: schemas.AdminRecordCreateRequest,
    db: Session = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    """기존 기록이 있으면 삭제 후 재생성(upsert). Payroll 월 전체 재계산."""
    db.query(AttendanceEvent).filter_by(
        user_id=payload.user_id, work_date=payload.work_date
    ).delete()
    db.flush()

    _add_event(db, payload.user_id, payload.work_date, EventType.CLOCK_IN, payload.check_in, note=payload.note)
    if payload.break_start:
        _add_event(db, payload.user_id, payload.work_date, EventType.BREAK_START, payload.break_start)
    if payload.break_end:
        _add_event(db, payload.user_id, payload.work_date, EventType.BREAK_END, payload.break_end)
    if payload.check_out:
        _add_event(db, payload.user_id, payload.work_date, EventType.CLOCK_OUT, payload.check_out)

    db.flush()
    AttendanceService.recalculate_payroll_for_month(
        db, payload.user_id, payload.work_date.year, payload.work_date.month
    )
    db.commit()
    return _build_summary(db, payload.user_id, payload.work_date)


@router.patch(
    "/admin/record/{user_id}/{work_date}",
    response_model=schemas.DailySummary,
    summary="[관리자] 근태 수정",
)
def admin_update_record(
    user_id: int,
    work_date: date,
    payload: schemas.AdminRecordUpdateRequest,
    db: Session = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    """기존 이벤트 전체 삭제 후 새 값으로 재생성. Payroll 월 전체 재계산."""
    existing = (
        db.query(AttendanceEvent)
        .filter_by(user_id=user_id, work_date=work_date)
        .first()
    )
    if not existing:
        raise HTTPException(status_code=404, detail="근태 기록을 찾을 수 없습니다.")

    db.query(AttendanceEvent).filter_by(user_id=user_id, work_date=work_date).delete()
    db.flush()

    _add_event(db, user_id, work_date, EventType.CLOCK_IN, payload.check_in, note=payload.note)
    if payload.break_start:
        _add_event(db, user_id, work_date, EventType.BREAK_START, payload.break_start)
    if payload.break_end:
        _add_event(db, user_id, work_date, EventType.BREAK_END, payload.break_end)
    if payload.check_out:
        _add_event(db, user_id, work_date, EventType.CLOCK_OUT, payload.check_out)

    db.flush()
    AttendanceService.recalculate_payroll_for_month(
        db, user_id, work_date.year, work_date.month
    )
    db.commit()
    return _build_summary(db, user_id, work_date)


# ════════════════════════════════════════════════════════
# 본인 근태 조회 API
# ════════════════════════════════════════════════════════

@router.get(
    "/my/monthly",
    response_model=schemas.MonthlyAttendanceResponse,
    summary="내 월별 근태 조회",
)
def my_monthly_attendance(
    year: int = Query(...),
    month: int = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    records = AttendanceService.get_monthly_attendance(
        db, year, month, current_user.id, ignore_position_filter=True
    )
    return schemas.MonthlyAttendanceResponse(
        records=[schemas.DailySummary(**r) for r in records],
        total=len(records),
    )


@router.delete(
    "/admin/record/{user_id}/{work_date}",
    status_code=204,
    summary="[관리자] 근태 삭제",
)
def admin_delete_record(
    user_id: int,
    work_date: date,
    db: Session = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    """해당 날짜 모든 이벤트 삭제 및 Payroll 월 전체 재계산."""
    deleted = (
        db.query(AttendanceEvent)
        .filter_by(user_id=user_id, work_date=work_date)
        .delete()
    )
    if not deleted:
        raise HTTPException(status_code=404, detail="근태 기록을 찾을 수 없습니다.")

    db.flush()
    AttendanceService.recalculate_payroll_for_month(
        db, user_id, work_date.year, work_date.month
    )
    db.commit()
