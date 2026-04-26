"""
급여 서비스
──────────────────────────────────────────────────
- 관리자: 전직원 급여 조회/수정/엑셀 다운로드
- 직원:  본인 급여 명세서 조회

급여 계산 기준 (한국 노동법):
  주간급여   = wage × day_hours
  야간급여   = wage × night_hours × 1.5  (야간가산 포함)
  주휴수당   = wage × weekly_allowance_hours
  연차수당   = wage × annual_leave_hours
  공휴일수당 = wage × holiday_hours × 1.5  (주간·야간 시간은 각각 day_hours·night_hours에도 포함)


공제 (연도별 InsuranceRate 기준):
  건강보험   = gross × health_insurance_rate / 100
  요양보험   = 건강보험료 × long_term_care_rate / 100
  고용보험   = gross × employment_insurance_rate / 100
  국민연금   = gross × national_pension_rate / 100
"""

from __future__ import annotations

import io
import math
from datetime import date, timedelta
from decimal import ROUND_DOWN, Decimal
from typing import List, Optional, Union

import openpyxl
from sqlalchemy.orm import Session

from app.modules.admin.models import Holiday, InsuranceRate
from app.modules.auth.models import PositionEnum, User
from app.modules.auth.services import decrypt_ssn
from app.modules.payroll.models import Payroll, PayrollPayDate
from app.modules.payroll.schemas import (
    PayrollAdminUpdateInput,
    PayrollPayResponse,
    PayrollResponse,
)
from app.modules.workstatus.models import AttendanceEvent, EventType
from app.utils.permission_utils import is_admin


class PayrollService:

    @staticmethod
    def get_payrolls(
        *,
        db: Session,
        user: User,
        year: int,
        month: Optional[int] = None,
    ) -> Union[List[PayrollResponse], PayrollPayResponse]:
        """급여 조회 (관리자: 전체, 직원: 본인)"""

        if is_admin(user):
            query = (
                db.query(Payroll)
                .join(User, Payroll.user_id == User.id)
                .filter(Payroll.year == year)
            )
            if month is not None:
                query = query.filter(Payroll.month == month)

            payrolls = query.order_by(User.name).all()
            return [PayrollService._to_admin_response(p, db) for p in payrolls]

        # 직원 본인
        payroll = (
            db.query(Payroll)
            .filter(
                Payroll.user_id == user.id,
                Payroll.year == year,
                Payroll.month == month,
            )
            .first()
        )
        if not payroll:
            return PayrollPayResponse()

        return PayrollService._to_user_pay_response(payroll, db)

    @staticmethod
    def update_payroll(
        *,
        db: Session,
        payroll_id: int,
        data: PayrollAdminUpdateInput,
    ) -> PayrollResponse:
        """관리자 급여 항목 수정 (총급여·총공제·실수령액 제외)"""
        payroll = db.get(Payroll, payroll_id)
        if not payroll:
            from fastapi import HTTPException
            raise HTTPException(status_code=404, detail="급여 기록을 찾을 수 없습니다.")

        update_fields = data.model_dump(exclude_none=True)
        for field, value in update_fields.items():
            setattr(payroll, field, value)

        db.commit()
        db.refresh(payroll)
        return PayrollService._to_admin_response(payroll, db)

    # ──────────────────────────────────────────────────────
    # 내부: 관리자 응답 생성
    # ──────────────────────────────────────────────────────
    @staticmethod
    def _to_admin_response(payroll: Payroll, db: Session) -> PayrollResponse:
        user = payroll.user

        w = payroll.wage

        day_wage = _ceil10(w * float(payroll.day_hours))
        night_wage = _ceil10(w * float(payroll.night_hours) * 1.5)
        weekly_allowance_pay = (
            payroll.weekly_allowance_pay
            if payroll.weekly_allowance_pay is not None
            else _ceil10(w * float(payroll.weekly_allowance_hours))
        )
        # ── 입사 당월 규칙 판단 ───────────────────────────────────
        hire_date = user.hire_date
        is_hire_month = (
            hire_date is not None
            and hire_date.year == payroll.year
            and hire_date.month == payroll.month
        )
        # 규칙1: 당월 입사이고 1일이 아니면 연차 미발생
        effective_leave_hours = (
            0.0
            if is_hire_month and hire_date.day != 1
            else float(payroll.annual_leave_hours)
        )
        # 규칙2: 당월 입사 크루는 건강·요양·국민연금 미발생
        is_hire_month_crew = is_hire_month and user.position == PositionEnum.crew

        annual_leave_count = payroll.annual_leave_count or 1
        annual_leave_pay = (
            payroll.annual_leave_pay
            if payroll.annual_leave_pay is not None
            else _ceil10(w * effective_leave_hours * annual_leave_count)
        )
        holiday_pay = _ceil10(w * float(payroll.holiday_hours) * 1.5)

        gross_pay = (
            day_wage
            + night_wage
            + weekly_allowance_pay
            + annual_leave_pay
            + holiday_pay
        )

        total_work_hours = (
            float(payroll.day_hours)
            + float(payroll.night_hours)
        )

        # 보험료 계산
        rate_year = _insurance_rate_year(payroll.year, payroll.month)
        rate = _get_insurance_rate(db, rate_year)
        health, care, employment, pension = _calc_insurance(
            gross_pay,
            payroll.insurance_health,
            payroll.insurance_care,
            payroll.insurance_employment,
            payroll.insurance_pension,
            rate,
        )
        # 규칙2 적용: 당월 입사 크루는 고용보험 외 공제 없음
        if is_hire_month_crew:
            health = Decimal("0")
            care = Decimal("0")
            pension = Decimal("0")

        total_deduction = health + care + employment + pension

        # 근무일수 집계
        total_work_days = _count_work_days(db, payroll.user_id, payroll.year, payroll.month)

        # SSN 복호화
        raw_ssn = None
        if user.ssn:
            try:
                raw_ssn = decrypt_ssn(user.ssn)
            except Exception:
                raw_ssn = user.ssn

        return PayrollResponse(
            payroll_id=payroll.id,
            user_id=user.id,
            year=payroll.year,
            month=payroll.month,
            # 인적 정보
            name=user.name,
            position=user.position.value if user.position else None,
            wage=payroll.wage,
            rrn=raw_ssn,
            join_date=user.hire_date,
            resign_date=user.retire_date,
            last_work_day=payroll.last_work_day,
            bank_name=user.bank_name,
            bank_account=user.account_number,
            email=user.email,
            # 근무 요약
            total_work_days=total_work_days,
            total_work_hours=round(total_work_hours, 2),
            avg_daily_hours=(
                round(total_work_hours / total_work_days, 2)
                if total_work_days > 0 else None
            ),
            # 시간 항목
            day_hours=float(payroll.day_hours),
            night_hours=float(payroll.night_hours),
            weekly_allowance_hours=float(payroll.weekly_allowance_hours),
            annual_leave_hours=effective_leave_hours,
            annual_leave_count=annual_leave_count,
            holiday_hours=float(payroll.holiday_hours),
            # 급여 항목
            day_wage=day_wage,
            night_wage=night_wage,
            weekly_allowance_pay=weekly_allowance_pay,
            annual_leave_pay=annual_leave_pay,
            holiday_pay=holiday_pay,
            gross_pay=gross_pay,
            # 공제
            insurance_health=int(health),
            insurance_care=int(care),
            insurance_employment=int(employment),
            insurance_pension=int(pension),
            total_deduction=int(total_deduction),
            net_pay=gross_pay - int(total_deduction),
        )

    # ──────────────────────────────────────────────────────
    # 내부: 직원 응답 생성
    # ──────────────────────────────────────────────────────
    @staticmethod
    def _to_user_pay_response(payroll: Payroll, db: Session) -> PayrollPayResponse:
        user = payroll.user
        pay_date = _get_pay_date(db, payroll.year, payroll.month)

        w = payroll.wage

        day_pay = _ceil10(w * float(payroll.day_hours))
        night_pay = _ceil10(w * float(payroll.night_hours) * 1.5)
        weekly_allowance_pay = (
            payroll.weekly_allowance_pay
            if payroll.weekly_allowance_pay is not None
            else _ceil10(w * float(payroll.weekly_allowance_hours))
        )
        # ── 입사 당월 규칙 판단 ───────────────────────────────────
        hire_date = user.hire_date
        is_hire_month = (
            hire_date is not None
            and hire_date.year == payroll.year
            and hire_date.month == payroll.month
        )
        effective_leave_hours = (
            0.0
            if is_hire_month and hire_date.day != 1
            else float(payroll.annual_leave_hours)
        )
        is_hire_month_crew = is_hire_month and user.position == PositionEnum.crew

        annual_leave_count = payroll.annual_leave_count or 1
        annual_leave_pay = _ceil10(w * effective_leave_hours * annual_leave_count)
        holiday_pay = _ceil10(w * float(payroll.holiday_hours) * 1.5)
        gross_pay = (
            day_pay
            + night_pay
            + weekly_allowance_pay
            + annual_leave_pay
            + holiday_pay
        )

        rate_year = _insurance_rate_year(payroll.year, payroll.month)
        rate = _get_insurance_rate(db, rate_year)
        health, care, employment, pension = _calc_insurance(
            gross_pay,
            payroll.insurance_health,
            payroll.insurance_care,
            payroll.insurance_employment,
            payroll.insurance_pension,
            rate,
        )
        # 규칙2 적용: 당월 입사 크루는 고용보험 외 공제 없음
        if is_hire_month_crew:
            health = Decimal("0")
            care = Decimal("0")
            pension = Decimal("0")

        total_deduction = health + care + employment + pension

        total_work_hours = (
            float(payroll.day_hours)
            + float(payroll.night_hours)
        )
        total_work_days = _count_work_days(db, payroll.user_id, payroll.year, payroll.month)

        return PayrollPayResponse(
            name=user.name,
            position=user.position.value if user.position else None,
            birth_date=user.birth_date,
            pay_date=pay_date,
            wage=payroll.wage,
            # 시간 항목
            total_work_days=total_work_days,
            total_work_hours=round(total_work_hours, 2),
            avg_daily_hours=(
                round(total_work_hours / total_work_days, 2)
                if total_work_days > 0 else None
            ),
            day_hours=float(payroll.day_hours),
            night_hours=float(payroll.night_hours),
            weekly_allowance_hours=float(payroll.weekly_allowance_hours),
            annual_leave_hours=effective_leave_hours,
            holiday_hours=float(payroll.holiday_hours),
            # 급여 항목
            day_wage=day_pay,
            night_wage=night_pay,
            weekly_allowance_pay=weekly_allowance_pay,
            annual_leave_pay=annual_leave_pay,
            holiday_pay=holiday_pay,
            gross_pay=gross_pay,
            # 공제
            insurance_health=int(health),
            insurance_care=int(care),
            insurance_employment=int(employment),
            insurance_pension=int(pension),
            total_deduction=int(total_deduction),
            net_pay=gross_pay - int(total_deduction),
        )

    # ──────────────────────────────────────────────────────
    # 엑셀 내보내기
    # ──────────────────────────────────────────────────────
    @staticmethod
    def export_to_excel(
        payrolls: List[PayrollResponse],
        year: int,
        month: int,
    ) -> io.BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{year}년{month}월 급여"

        # 직급 제거 후 30개 열 (A~AD)
        headers = [
            "이름", "주민등록번호", "시급",                          # A B C
            "은행명", "계좌번호",                                    # D E
            "입사일", "퇴사일", "마지막근무일",                      # F G H
            "근무일수", "총근무시간", "일평균시간",                   # I J K
            "주간시간", "야간시간", "주휴시간", "연차시간",            # L M N O
            "공휴일시간",                                            # P
            "주간급여", "야간급여", "주휴수당", "연차수당",            # Q R S T
            "공휴일수당",                                            # U
            "급여총액",                                              # V
            "건강보험", "요양보험", "고용보험", "국민연금",            # W X Y Z
            "공제계", "실수령액",                                    # AA AB
        ]
        ws.append(headers)

        # ── 스타일 정의 ─────────────────────────────────────
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
        from openpyxl.utils import get_column_letter

        header_fill    = PatternFill(start_color="1A0F3C", end_color="1A0F3C", fill_type="solid")
        header_font    = Font(color="FFFFFF", bold=True, size=10)
        emphasis_fill  = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")  # 연보라
        emphasis_font  = Font(bold=True, size=10, color="1E1B4B")
        sum_fill       = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        sum_font       = Font(bold=True, size=10)
        zebra_fill     = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        data_font      = Font(size=10)
        thin           = Side(style="thin", color="D1D5DB")
        border         = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 열 분류 (1-based)
        EMPHASIS_COLS  = {22, 27, 28}           # 급여총액, 공제계, 실수령액
        MONEY_COLS     = {3, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28}  # 콤마 + 원
        HOURS_COLS     = {10, 11, 12, 13, 14, 15, 16}                           # 소수점 2자리
        RIGHT_COLS     = MONEY_COLS | HOURS_COLS | {9}                          # 우측 정렬
        SUM_COL_START  = 9
        SUM_COL_END    = 28

        FMT_MONEY = '#,##0'
        FMT_HOURS = '0.00'
        FMT_DAYS  = '0'

        # ── 헤더 행 ─────────────────────────────────────────
        ws.row_dimensions[1].height = 30
        for col_idx, cell in enumerate(ws[1], 1):
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = border

        # ── 데이터 행 ────────────────────────────────────────
        for row_num, p in enumerate(payrolls, start=2):
            ws.append([
                p.name, p.rrn or "", p.wage,
                p.bank_name or "", p.bank_account or "",
                str(p.join_date) if p.join_date else "",
                str(p.resign_date) if p.resign_date else "",
                str(p.last_work_day) if p.last_work_day else "",
                p.total_work_days or 0,
                p.total_work_hours or 0,
                p.avg_daily_hours or 0,
                p.day_hours or 0, p.night_hours or 0,
                p.weekly_allowance_hours or 0, p.annual_leave_hours or 0,
                p.holiday_hours or 0,
                p.day_wage or 0, p.night_wage or 0,
                p.weekly_allowance_pay or 0, p.annual_leave_pay or 0,
                p.holiday_pay or 0,
                p.gross_pay or 0,
                p.insurance_health or 0, p.insurance_care or 0,
                p.insurance_employment or 0, p.insurance_pension or 0,
                p.total_deduction or 0, p.net_pay or 0,
            ])
            ws.row_dimensions[row_num].height = 22
            is_zebra = (row_num % 2 == 0)
            for col_idx, cell in enumerate(ws[row_num], 1):
                cell.border = border
                cell.font   = data_font
                if col_idx in EMPHASIS_COLS:
                    cell.fill = emphasis_fill
                    cell.font = emphasis_font
                elif is_zebra:
                    cell.fill = zebra_fill
                # 숫자 서식
                if col_idx in MONEY_COLS:
                    cell.number_format = FMT_MONEY
                elif col_idx in HOURS_COLS:
                    cell.number_format = FMT_HOURS
                elif col_idx == 9:
                    cell.number_format = FMT_DAYS
                # 정렬
                if col_idx in RIGHT_COLS:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_idx in {1, 4, 5}:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── 합계 행 ─────────────────────────────────────────
        data_start = 2
        data_end   = ws.max_row
        sum_row: list = ["합계", "", "", "", "", "", "", ""]
        for col_idx in range(SUM_COL_START, SUM_COL_END + 1):
            col_letter = get_column_letter(col_idx)
            sum_row.append(f"=SUM({col_letter}{data_start}:{col_letter}{data_end})")
        ws.append(sum_row)

        sum_row_idx = ws.max_row
        ws.row_dimensions[sum_row_idx].height = 24
        for col_idx, cell in enumerate(ws[sum_row_idx], 1):
            cell.border = border
            if col_idx in EMPHASIS_COLS:
                cell.fill = emphasis_fill
                cell.font = emphasis_font
            else:
                cell.fill = sum_fill
                cell.font = sum_font
            if col_idx in MONEY_COLS:
                cell.number_format = FMT_MONEY
            elif col_idx in HOURS_COLS:
                cell.number_format = FMT_HOURS
            elif col_idx == 9:
                cell.number_format = FMT_DAYS
            if col_idx in RIGHT_COLS:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in {1, 4, 5}:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── 열 너비 고정 ─────────────────────────────────────
        col_widths = {
            1: 10,   # 이름
            2: 16,   # 주민등록번호
            3: 10,   # 시급
            4: 12,   # 은행명
            5: 16,   # 계좌번호
            6: 12,   # 입사일
            7: 12,   # 퇴사일
            8: 13,   # 마지막근무일
            9: 9,    # 근무일수
            10: 11,  # 총근무시간
            11: 11,  # 일평균시간
            12: 9, 13: 9, 14: 9, 15: 9, 16: 10,  # 시간 열
            17: 12, 18: 12, 19: 12, 20: 12, 21: 12,  # 급여 열
            22: 14,  # 급여총액
            23: 11, 24: 11, 25: 11, 26: 11,  # 보험
            27: 11,  # 공제계
            28: 14,  # 실수령액
        }
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 틀 고정 (헤더 + 이름 열)
        ws.freeze_panes = "B2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ──────────────────────────────────────────────────────
    # 대량 업로드 양식 생성
    # ──────────────────────────────────────────────────────
    @staticmethod
    def generate_bulk_template(db: Session) -> io.BytesIO:
        """대량 업로드용 엑셀 양식 생성 (직원 목록 시트 포함)"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── 입력 시트 ────────────────────────────────────────
        ws = wb.active
        ws.title = "급여_입력"

        headers = [
            "직원ID",        # A — 필수, users.id
            "이름",          # B — 참고용
            "연도",          # C — 필수
            "월",            # D — 필수
            "시급",          # E
            "주간시간",       # F
            "야간시간",       # G
            "주휴수당",        # H — 비워두면 wage×주휴시간 자동계산
            "연차시간",       # I
            "공휴일시간",     # J
            "연차수당직접입력",  # K — 비워두면 자동계산
            "건강보험",       # L — 0 또는 빈칸이면 자동계산
            "요양보험",       # M
            "고용보험",       # N
            "국민연금",       # O
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color="1A0F3C", end_color="1A0F3C", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        required_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        thin = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        REQUIRED_COLS = {1, 2, 3, 4}  # 직원ID, 이름, 연도, 월

        ws.row_dimensions[1].height = 28
        for col_idx, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # 안내 행 (2행)
        ws.append([
            "← 직원목록 시트 참고",
            "",
            "예: 2026",
            "예: 1",
            "",
            "",
            "",
            "비워두면 자동",
            "",
            "",
            "비워두면 자동",
            "비워두면 자동",
            "비워두면 자동",
            "비워두면 자동",
            "비워두면 자동",
        ])
        guide_row = ws[2]
        guide_font = Font(color="9CA3AF", italic=True, size=9)
        for cell in guide_row:
            cell.font = guide_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        col_widths = {
            1: 10, 2: 12, 3: 8, 4: 6, 5: 10,
            6: 10, 7: 10, 8: 10, 9: 10, 10: 12,
            11: 16, 12: 12, 13: 10, 14: 10, 15: 10,
        }
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A3"

        # ── 직원 목록 시트 ───────────────────────────────────
        ws2 = wb.create_sheet("직원목록")
        ws2.append(["직원ID", "이름", "직급", "입사일", "상태"])
        for col_idx, cell in enumerate(ws2[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        users = (
            db.query(User)
            .filter(User.position.notin_(["system"]))
            .order_by(User.name)
            .all()
        )
        for u in users:
            pos = u.position.value if hasattr(u.position, "value") else str(u.position)
            status_str = "재직" if u.is_active else "퇴직"
            ws2.append([u.id, u.name, pos, str(u.hire_date) if u.hire_date else "", status_str])

        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws2.column_dimensions["A"].width = 10
        ws2.column_dimensions["B"].width = 12
        ws2.column_dimensions["C"].width = 10
        ws2.column_dimensions["D"].width = 14
        ws2.column_dimensions["E"].width = 8

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ──────────────────────────────────────────────────────
    # 대량 업로드 처리
    # ──────────────────────────────────────────────────────
    @staticmethod
    def bulk_upload(db: Session, file_bytes: bytes) -> dict:
        """
        엑셀 파일을 파싱해 Payroll 레코드를 upsert합니다.
        반환: {"inserted": int, "updated": int, "errors": list[str]}
        """
        from decimal import Decimal as D
        from app.modules.payroll.models import Payroll

        wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        inserted = 0
        updated = 0
        errors: list[str] = []

        def _to_decimal(val, default="0.00") -> D:
            try:
                return D(str(val)).quantize(D("0.01")) if val not in (None, "") else D(default)
            except Exception:
                return D(default)

        def _to_int(val, default=0) -> int:
            try:
                return int(float(str(val))) if val not in (None, "") else default
            except Exception:
                return default

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 완전히 빈 행 스킵
            if all(v is None or str(v).strip() == "" for v in row[:4]):
                continue

            # 첫 번째 열이 숫자가 아니면 헤더/안내 행으로 간주하고 조용히 스킵
            try:
                float(str(row[0]))
            except (ValueError, TypeError):
                continue

            try:
                user_id = _to_int(row[0])
                year = _to_int(row[2])
                month = _to_int(row[3])

                if user_id <= 0 or year <= 0 or month < 1 or month > 12:
                    errors.append(f"행 {row_num}: 직원ID({row[0]}), 연도({row[2]}), 월({row[3]}) 값이 유효하지 않습니다.")
                    continue

                user = db.get(User, user_id)
                if not user:
                    errors.append(f"행 {row_num}: 직원ID {user_id}에 해당하는 직원이 없습니다.")
                    continue

                wage = _to_int(row[4]) or user.wage or 0
                day_hours = _to_decimal(row[5])
                night_hours = _to_decimal(row[6])
                weekly_allowance_pay_raw = row[7]
                weekly_allowance_pay_val = _to_int(weekly_allowance_pay_raw) if weekly_allowance_pay_raw not in (None, "") else None
                annual_leave_hours = _to_decimal(row[8])
                holiday_hours = _to_decimal(row[9])
                annual_leave_pay_raw = row[10]
                annual_leave_pay_val = _to_int(annual_leave_pay_raw) if annual_leave_pay_raw not in (None, "") else None
                insurance_health = _to_int(row[11])
                insurance_care = _to_int(row[12])
                insurance_employment = _to_int(row[13])
                insurance_pension = _to_int(row[14])

                existing = (
                    db.query(Payroll)
                    .filter(Payroll.user_id == user_id, Payroll.year == year, Payroll.month == month)
                    .first()
                )

                if existing:
                    existing.wage = wage
                    existing.day_hours = day_hours
                    existing.night_hours = night_hours
                    existing.weekly_allowance_pay = weekly_allowance_pay_val
                    existing.annual_leave_hours = annual_leave_hours
                    existing.holiday_hours = holiday_hours
                    existing.annual_leave_pay = annual_leave_pay_val
                    existing.insurance_health = insurance_health
                    existing.insurance_care = insurance_care
                    existing.insurance_employment = insurance_employment
                    existing.insurance_pension = insurance_pension
                    updated += 1
                else:
                    payroll = Payroll(
                        user_id=user_id,
                        year=year,
                        month=month,
                        wage=wage,
                        day_hours=day_hours,
                        night_hours=night_hours,
                        weekly_allowance_pay=weekly_allowance_pay_val,
                        annual_leave_hours=annual_leave_hours,
                        holiday_hours=holiday_hours,
                        annual_leave_pay=annual_leave_pay_val,
                        insurance_health=insurance_health,
                        insurance_care=insurance_care,
                        insurance_employment=insurance_employment,
                        insurance_pension=insurance_pension,
                    )
                    db.add(payroll)
                    inserted += 1

            except Exception as e:
                errors.append(f"행 {row_num}: 처리 중 오류 — {str(e)}")

        db.commit()
        return {"inserted": inserted, "updated": updated, "errors": errors}


# ──────────────────────────────────────────────────────
# 유틸 함수
# ──────────────────────────────────────────────────────

def _ceil10(value: float) -> int:
    """급여항목: 1의 자리 올림 처리 (13 → 20)"""
    return math.ceil(value / 10) * 10


def _insurance_rate_year(year: int, month: int) -> int:
    return year


def _get_insurance_rate(db: Session, rate_year: int) -> Optional[InsuranceRate]:
    return db.query(InsuranceRate).filter(InsuranceRate.year == rate_year).first()


def _calc_insurance(
    gross_pay: int,
    stored_health: int,
    stored_care: int,
    stored_employment: int,
    stored_pension: int,
    rate: Optional[InsuranceRate],
) -> tuple[Decimal, Decimal, Decimal, Decimal]:
    """
    공제 계산 — 저장된 값이 0이면 요율로 자동 계산
    (관리자가 직접 수정한 경우 저장값 우선)
    """
    gp = Decimal(gross_pay)

    health = Decimal(stored_health or 0)
    care = Decimal(stored_care or 0)
    employment = Decimal(stored_employment or 0)
    pension = Decimal(stored_pension or 0)

    if rate:
        if health == 0:
            health = (
                gp * rate.health_insurance_rate / Decimal("100")
            ).quantize(Decimal("1E1"), rounding=ROUND_DOWN)

        if care == 0:
            care = (
                health * rate.long_term_care_rate / Decimal("100")
            ).quantize(Decimal("1E1"), rounding=ROUND_DOWN)

        if employment == 0:
            employment = (
                gp * rate.employment_insurance_rate / Decimal("100")
            ).quantize(Decimal("1E1"), rounding=ROUND_DOWN)

        if pension == 0:
            pension = (
                gp * rate.national_pension_rate / Decimal("100")
            ).quantize(Decimal("1E1"), rounding=ROUND_DOWN)

    return health, care, employment, pension


def _count_work_days(db: Session, user_id: int, year: int, month: int) -> int:
    """해당 월 CLOCK_OUT 이벤트 기준 근무일 수"""
    month_start = date(year, month, 1)
    month_end = (
        date(year, month + 1, 1) if month < 12 else date(year + 1, 1, 1)
    )
    return (
        db.query(AttendanceEvent.work_date)
        .filter(
            AttendanceEvent.user_id == user_id,
            AttendanceEvent.event_type == EventType.CLOCK_OUT,
            AttendanceEvent.work_date >= month_start,
            AttendanceEvent.work_date < month_end,
        )
        .distinct()
        .count()
    )


def _get_pay_date(db: Session, year: int, month: int) -> Optional[date]:
    pay_date = (
        db.query(PayrollPayDate)
        .filter(PayrollPayDate.year == year, PayrollPayDate.month == month)
        .first()
    )
    return pay_date.pay_date if pay_date else None


def calculate_auto_pay_date(db: Session, year: int, month: int, payment_day: int = 10) -> date:
    """해당월 급여는 다음달 payment_day일에 지급. 주말/공휴일이면 직전 평일을 반환한다."""
    import calendar
    # 다음달 계산
    if month == 12:
        pay_year, pay_month = year + 1, 1
    else:
        pay_year, pay_month = year, month + 1

    last_day = calendar.monthrange(pay_year, pay_month)[1]
    day = min(payment_day, last_day)
    target = date(pay_year, pay_month, day)

    holidays = {
        h.date for h in db.query(Holiday).filter(
            Holiday.date >= date(pay_year, pay_month, 1),
            Holiday.date <= date(pay_year, pay_month, last_day),
        ).all()
    }

    while target.weekday() >= 5 or target in holidays:
        target -= timedelta(days=1)

    return target


def upsert_pay_date(db: Session, year: int, month: int, pay_date: date) -> PayrollPayDate:
    record = (
        db.query(PayrollPayDate)
        .filter(PayrollPayDate.year == year, PayrollPayDate.month == month)
        .first()
    )
    if record:
        record.pay_date = pay_date
    else:
        record = PayrollPayDate(year=year, month=month, pay_date=pay_date)
        db.add(record)
    db.commit()
    db.refresh(record)
    return record
